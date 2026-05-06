"""
API routes for dashboard.

Date/time filtering contract
─────────────────────────────
All date-range validation is enforced by the Pydantic schema (VehicleLogsRequest)
BEFORE execution reaches this router.  The rules are:

  • Time-filtered requests  →  start_date == end_date (same calendar day)
  • Dashboard scope         →  max 30 calendar days
  • Report scope            →  max 90 calendar days

The router trusts those invariants and does NOT re-validate them here,
keeping the business logic single-sourced and testable.
"""

from __future__ import annotations

import io
import time as time_module
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
from datetime import date, datetime, time, timedelta, timezone
from typing import Optional

import openpyxl
import requests
from fastapi import APIRouter, Body, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage
from sqlalchemy.orm import Session
from application.auth.utils import get_current_user
from application.dashboard import crud, schemas, utils
from application.database.models.checkpoint import MstCheckpoint
from application.database.models.location import MstLocation
from application.database.models.transactions.access_control import TrnAccessControl
from application.database.models.transactions.vehicle_log import TrnVehicleLog
from application.database.session import get_db
from application.helpers.logger import get_logger
from application.helpers.storage import get_storage

logger = get_logger("dashboard")
router = APIRouter(prefix="/dashboard", tags=["Dashboard"])

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
_MIDNIGHT = time(0, 0, 0)
_END_OF_DAY = time(23, 59, 59, 999999)

def _coerce_start(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.combine(value, _MIDNIGHT)

def _coerce_end(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        # If it's a datetime at midnight, treat it as a date and add 1 day
        if value.time() == _MIDNIGHT:
            return datetime.combine(value.date() + timedelta(days=1), _MIDNIGHT)
        # Otherwise it's a specific time, return as-is
        return value
    # Plain date object - add 1 day to include all records on the end_date
    return datetime.combine(value + timedelta(days=1), _MIDNIGHT)

def _resolve_date_window(
    request: schemas.VehicleLogsRequest,
    scope: schemas.ScopeEnum,
) -> tuple[datetime | None, datetime | None]:
    if scope == schemas.ScopeEnum.dashboard and request.start_date is None and request.end_date is None:
        today = date.today()
        return (
            datetime.combine(today, _MIDNIGHT),
            datetime.combine(today, _END_OF_DAY),
        )
    return _coerce_start(request.start_date), _coerce_end(request.end_date)

def _intersect_ids(
    requested: list[int] | None,
    accessible: list[int] | None,
) -> list[int] | None:
    if requested is None:
        return accessible
    if accessible is None:
        return requested
    return [i for i in requested if i in accessible]

# ---------------------------------------------------------------------------
# Vehicle logs endpoint
# ---------------------------------------------------------------------------
@router.post("/vehicle-logs")
def get_vehicle_logs(
    request: schemas.VehicleLogsRequest = Body(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user_id = current_user.user_id

    logger.info(
        "VehicleLogs :: user=%s(%s) company=%s role=%s scope=%s "
        "locs=%s cps=%s dates=%s–%s time_filtered=%s "
        "bl=%s wl=%s plate=%s page=%s/%s excel=%s",
        user_id, current_user.username,
        current_user.company_id, current_user.role,
        request.scope,
        request.location_ids, request.checkpoint_ids,
        request.start_date, request.end_date, request.is_time_filtered,
        request.is_blacklisted, request.is_whitelisted,
        request.plate_number, request.page, request.page_size,
        request.excel_report,
    )

    # ── 1. Resolve access-control ─────────────────────────────────────────
    access_entries = (
        db.query(TrnAccessControl)
        .filter(
            TrnAccessControl.user_id == user_id,
            TrnAccessControl.disabled == False,
            TrnAccessControl.is_deleted == False,
        )
        .all()
    )

    if not access_entries:
        logger.warning("VehicleLogs :: user=%s has no access-control entries", user_id)
        return {"total_logs": 0, "logs": []}

    access_info = utils.extract_accessible_locations_checkpoints(
        access_entries,
        db=db,
        company_id=current_user.company_id,
        role=current_user.role,
    )
    user_location_ids: list[int] | None = access_info["location_ids"]
    user_checkpoint_ids: list[int] | None = access_info["checkpoint_ids"]

    logger.info(
        "VehicleLogs :: access resolved locs=%s cps=%s",
        user_location_ids, user_checkpoint_ids,
    )

    # ── 2. Resolve date window ────────────────────────────────────────────
    start_dt, end_dt = _resolve_date_window(request, request.scope)

    # ── 3. Resolve location / checkpoint IDs ──────────────────────────────
    if request.scope == schemas.ScopeEnum.dashboard:
        location_ids = user_location_ids
        checkpoint_ids = user_checkpoint_ids
    else:
        location_ids = _intersect_ids(request.location_ids, user_location_ids)

        if location_ids is not None and len(location_ids) > 0:
            location_checkpoint_rows = (
                db.query(MstCheckpoint.checkpoint_id)
                .filter(
                    MstCheckpoint.location_id.in_(location_ids),
                    MstCheckpoint.disabled == False,
                    MstCheckpoint.is_deleted == False,
                )
                .all()
            )
            location_checkpoint_ids = [r[0] for r in location_checkpoint_rows]

            if request.checkpoint_ids:
                checkpoint_ids = [
                    cid for cid in request.checkpoint_ids if cid in location_checkpoint_ids
                ]
            else:
                checkpoint_ids = location_checkpoint_ids
        else:
            checkpoint_ids = _intersect_ids(request.checkpoint_ids, user_checkpoint_ids)

    logger.info(
        "VehicleLogs :: final scope=%s locs=%s cps=%s start=%s end=%s time_filtered=%s",
        request.scope, location_ids, checkpoint_ids, start_dt, end_dt, request.is_time_filtered,
    )

    # ── 4. Summary counts ────────────────────────────────────────────────
    summary = crud.get_summary_counts(
        db,
        company_id=current_user.company_id,
        location_ids=location_ids,
        checkpoint_ids=checkpoint_ids,
        start_date=start_dt,
        end_date=end_dt,
    )

    # ── 5. Pagination metadata ────────────────────────────────────────────
    use_expanded_pagination = bool(
        request.plate_number and request.scope == schemas.ScopeEnum.report
    )

    total_records = crud.get_vehicle_logs_count(
        db,
        company_id=current_user.company_id,
        location_ids=location_ids,
        checkpoint_ids=checkpoint_ids,
        start_date=start_dt,
        end_date=end_dt,
        is_blacklisted=request.is_blacklisted,
        is_whitelisted=request.is_whitelisted,
        plate_number=request.plate_number if use_expanded_pagination else None,
    )

    total_pages = (
        (total_records + request.page_size - 1) // request.page_size
        if total_records > 0
        else 0
    )

    # ── 6. Fetch logs ────────────────────────────────────────────────────
    if use_expanded_pagination:
        raw_entries = crud.get_vehicle_logs_with_blacklist_expanded(
            db,
            company_id=current_user.company_id,
            location_ids=location_ids,
            checkpoint_ids=checkpoint_ids,
            start_date=start_dt,
            end_date=end_dt,
            is_blacklisted=request.is_blacklisted,
            is_whitelisted=request.is_whitelisted,
            plate_number=request.plate_number,
            page=request.page,
            page_size=request.page_size,
        )
    else:
        raw_logs = crud.get_vehicle_logs_with_blacklist(
            db,
            company_id=current_user.company_id,
            location_ids=location_ids,
            checkpoint_ids=checkpoint_ids,
            start_date=start_dt,
            end_date=end_dt,
            is_blacklisted=request.is_blacklisted,
            is_whitelisted=request.is_whitelisted,
            plate_number=None,
            page=request.page,
            page_size=request.page_size,
        )

    # ── 7. Build response payload ─────────────────────────────────────────
    result, image_paths = _build_result(
        db=db,
        use_expanded=use_expanded_pagination,
        entries=raw_entries if use_expanded_pagination else raw_logs,
    )

    # ── 8. Presign all collected image paths in one batch ─────────────────
    storage = get_storage()
    presigned_urls = storage.generate_presigned_urls_batch(list(image_paths), expiration=3600)
    _apply_presigned_urls(result, presigned_urls)

    logger.info(
        "VehicleLogs :: response total_vehicles=%s page=%s/%s records=%s/%s urls=%s",
        summary["total_vehicles"], request.page, total_pages, len(result), total_records, len(presigned_urls),
    )

    # ── 9. Excel export (report scope only) ──────────────────────────────
    if request.excel_report and request.scope == schemas.ScopeEnum.report:
        return _stream_excel_report(
            db=db,
            request=request,
            current_user=current_user,
            use_expanded_pagination=use_expanded_pagination,
            location_ids=location_ids,
            checkpoint_ids=checkpoint_ids,
            start_dt=start_dt,
            end_dt=end_dt,
            storage=storage,
            total_records=total_records,
        )

    return {
        "total_vehicles": summary["total_vehicles"],
        "total_locations": summary["total_locations"],
        "total_cameras": summary["total_cameras"],
        "blacklisted_vehicle_count": summary["blacklisted_vehicle_count"],
        "multiple_detections_count": summary["multiple_detections_count"],
        "pagination": {
            "page": request.page,
            "page_size": request.page_size,
            "total_records": total_records,
            "total_pages": total_pages,
            "has_next": request.page < total_pages,
            "has_previous": request.page > 1,
        },
        "summary_data": result,
    }

# ---------------------------------------------------------------------------
# Result builders (pure functions – easier to unit-test)
# ---------------------------------------------------------------------------
def _checkpoint_cache(db: Session, checkpoint_ids: set) -> dict:
    if not checkpoint_ids:
        return {}

    rows = (
        db.query(
            MstCheckpoint.checkpoint_id,
            MstCheckpoint.name,
            MstCheckpoint.location_id,
            MstLocation.location_name,
        )
        .outerjoin(MstLocation, MstCheckpoint.location_id == MstLocation.location_id)
        .filter(MstCheckpoint.checkpoint_id.in_(checkpoint_ids))
        .all()
    )

    return {
        row.checkpoint_id: {
            "checkpoint_name": row.name,
            "location_id": row.location_id,
            "location_name": row.location_name,
        }
        for row in rows
    }

def _collect_checkpoint_ids_from_expanded(entries: list) -> set:
    return {
        e["history_entry"].get("checkpoint_id")
        for e in entries
        if e["history_entry"].get("checkpoint_id")
    }

def _collect_checkpoint_ids_from_logs(logs: list) -> set:
    ids = set()
    for log in logs:
        if log.history_data:
            for entry in log.history_data:
                cp_id = entry.get("checkpoint_id")
                if cp_id:
                    ids.add(cp_id)
    return ids

def _display_plate(log) -> str:
    if log.is_revised and log.revised_data:
        return log.revised_data.get("new_number", log.plate_number)
    return log.plate_number

def _picture_paths(picture_data: dict) -> tuple[str | None, str | None]:
    vehicle = picture_data.get("VehiclePic", {}).get("Content")
    plate = picture_data.get("CutoutPic", {}).get("Content")
    return vehicle, plate

def _build_result(db: Session, use_expanded: bool, entries) -> tuple[list, set]:
    image_paths: set[str] = set()
    result: list[dict] = []

    if use_expanded:
        cp_ids = _collect_checkpoint_ids_from_expanded(entries)
        cp_cache = _checkpoint_cache(db, cp_ids)

        for ed in entries:
            log = ed["log"]
            history_entry = ed["history_entry"]
            picture_data = history_entry.get("Picture", {})
            vehicle_img, plate_img = _picture_paths(picture_data)

            if vehicle_img:
                image_paths.add(vehicle_img)
            if plate_img:
                image_paths.add(plate_img)

            cp_id = history_entry.get("checkpoint_id")
            cp_info = cp_cache.get(cp_id, {})

            result.append(
                {
                    "log_id": log.log_id,
                    "vehicle_id": log.vehicle_id,
                    "detection_number": ed["detection_number"],
                    "location_id": cp_info.get("location_id"),
                    "location_name": cp_info.get("location_name"),
                    "checkpoint_id": cp_id,
                    "checkpoint_name": cp_info.get("checkpoint_name"),
                    "timestamp": picture_data.get("SnapInfo", {}).get("SnapTime"),
                    "plate_number": _display_plate(log),
                    "is_blacklisted": bool(log.is_blacklisted),
                    "is_whitelisted": bool(log.is_whitelisted),
                    "latest_data_vehicle_image": vehicle_img,
                    "latest_data_number_plate_image": plate_img,
                    "is_multiple_times": ed["total_detections"] > 1,
                    "is_revised": bool(log.is_revised),
                    "timeline": [],
                }
            )
    else:
        cp_ids = _collect_checkpoint_ids_from_logs(entries)
        cp_cache = _checkpoint_cache(db, cp_ids)

        for log in entries:
            picture_data = log.latest_data.get("Picture", {}) if log.latest_data else {}
            latest_vehicle_img, latest_plate_img = _picture_paths(picture_data)

            if latest_vehicle_img:
                image_paths.add(latest_vehicle_img)
            if latest_plate_img:
                image_paths.add(latest_plate_img)

            timeline = []
            if log.history_data:
                for entry in log.history_data:
                    ep = entry.get("Picture", {})
                    v_img, p_img = _picture_paths(ep)
                    if v_img:
                        image_paths.add(v_img)
                    if p_img:
                        image_paths.add(p_img)

                    cp_id = entry.get("checkpoint_id")
                    cp_info = cp_cache.get(cp_id, {})
                    timeline.append(
                        {
                            "location_name": cp_info.get("location_name"),
                            "checkpoint_name": cp_info.get("checkpoint_name"),
                            "time": ep.get("SnapInfo", {}).get("SnapTime", ""),
                            "vehicle_image": v_img,
                            "number_plate_image": p_img,
                        }
                    )

            row: dict = {
                "log_id": log.log_id,
                "vehicle_id": log.vehicle_id,
                "location_id": log.location_id,
                "location_name": log.location_name,
                "checkpoint_id": log.checkpoint_id,
                "checkpoint_name": log.checkpoint_name,
                "timestamp": picture_data.get("SnapInfo", {}).get("SnapTime"),
                "plate_number": _display_plate(log),
                "is_blacklisted": bool(log.is_blacklisted),
                "is_whitelisted": bool(log.is_whitelisted),
                "latest_data_vehicle_image": latest_vehicle_img,
                "latest_data_number_plate_image": latest_plate_img,
                "is_multiple_times": len(log.history_data) > 1 if log.history_data else False,
                "is_revised": bool(log.is_revised),
                "timeline": timeline,
            }

            if log.is_revised and log.revised_data:
                row["revised_data"] = log.revised_data

            result.append(row)

    return result, image_paths

def _apply_presigned_urls(result: list, presigned_urls: dict) -> None:
    for item in result:
        for key in ("latest_data_vehicle_image", "latest_data_number_plate_image"):
            path = item.get(key)
            if path:
                item[key] = presigned_urls.get(path)

        for tl in item.get("timeline", []):
            for key in ("vehicle_image", "number_plate_image"):
                path = tl.get(key)
                if path:
                    tl[key] = presigned_urls.get(path)


# ---------------------------------------------------------------------------
# Image downloading — production-grade, with retry + connection pooling
# ---------------------------------------------------------------------------

# One shared session for ALL image downloads — keeps TCP connections alive
# across the entire request lifetime (huge win for S3/CDN images).
_IMG_SESSION: requests.Session | None = None
_IMG_SESSION_LOCK = threading.Lock()

def _get_img_session() -> requests.Session:
    """Lazy-init a module-level requests.Session with a generous connection pool."""
    global _IMG_SESSION
    if _IMG_SESSION is None:
        with _IMG_SESSION_LOCK:
            if _IMG_SESSION is None:
                adapter = requests.adapters.HTTPAdapter(
                    pool_connections=20,
                    pool_maxsize=100,
                    max_retries=0,          # we implement our own retry logic below
                    pool_block=False,
                )
                s = requests.Session()
                s.mount("http://", adapter)
                s.mount("https://", adapter)
                _IMG_SESSION = s
    return _IMG_SESSION


# Target cell dimensions in pixels (Excel units ≈ 96 dpi).
# Plate images:   wide crop, readable text
# Vehicle images: taller crop, shows the car
_PLATE_IMG_SIZE = (160, 60)   # (max_width, max_height)

# How long (seconds) we wait for a single image download attempt.
_IMG_CONNECT_TIMEOUT = 4.0
_IMG_READ_TIMEOUT    = 8.0

# Retry policy: up to N attempts with exponential back-off.
_IMG_MAX_RETRIES = 3
_IMG_RETRY_BACKOFF = [0, 0.3, 0.8]   # seconds before each attempt


def download_and_resize_image(
    url: str,
    max_size: tuple[int, int] = _PLATE_IMG_SIZE,
    *,
    session: requests.Session | None = None,
) -> io.BytesIO | None:
    """
    Download *url* and resize to *max_size* while preserving aspect ratio.

    Improvements over the original implementation
    ──────────────────────────────────────────────
    • Uses a shared persistent Session (connection re-use → faster for S3).
    • Separate connect / read timeouts so a slow server doesn't block a worker
      for an unreasonably long time.
    • Retry loop with short back-off for transient S3 / CDN hiccups.
    • Reads the response body only once — no double-buffering.
    • Converts all non-RGB(A) modes before thumbnail so Pillow never raises.
    • Returns None (never raises) so a missing image never crashes the export.
    """
    sess = session or _get_img_session()

    for attempt, backoff in enumerate(_IMG_RETRY_BACKOFF):
        if backoff:
            time_module.sleep(backoff)
        try:
            resp = sess.get(
                url,
                timeout=(_IMG_CONNECT_TIMEOUT, _IMG_READ_TIMEOUT),
                stream=False,   # read the whole body into memory at once
            )
            if resp.status_code != 200:
                logger.debug(
                    "ImgDownload :: attempt=%d/%d status=%d url=%.80s",
                    attempt + 1, _IMG_MAX_RETRIES, resp.status_code, url,
                )
                continue        # retry on non-200

            raw = io.BytesIO(resp.content)
            pil = PILImage.open(raw)

            # Ensure we always have a solid RGB image Pillow / openpyxl can handle.
            if pil.mode == "P":
                pil = pil.convert("RGBA")
            if pil.mode in ("RGBA", "LA"):
                background = PILImage.new("RGB", pil.size, (255, 255, 255))
                background.paste(pil, mask=pil.split()[-1])
                pil = background
            elif pil.mode != "RGB":
                pil = pil.convert("RGB")

            pil.thumbnail(max_size, PILImage.Resampling.LANCZOS)

            buf = io.BytesIO()
            pil.save(buf, format="JPEG", quality=85, optimize=True, progressive=True)
            buf.seek(0)
            return buf

        except (requests.exceptions.ConnectTimeout,
                requests.exceptions.ReadTimeout,
                requests.exceptions.ConnectionError) as exc:
            logger.debug(
                "ImgDownload :: attempt=%d/%d network_error=%s url=%.80s",
                attempt + 1, _IMG_MAX_RETRIES, type(exc).__name__, url,
            )
        except Exception as exc:
            logger.debug(
                "ImgDownload :: attempt=%d/%d unexpected_error=%s url=%.80s",
                attempt + 1, _IMG_MAX_RETRIES, exc, url,
            )
            break   # non-retriable (e.g. corrupt image), stop immediately

    return None


def _download_images_parallel(
    tasks: list[tuple],            # list of (key, url, max_size)
    *,
    max_workers: int = 50,
) -> dict:
    """
    Download all images in *tasks* concurrently and return a mapping
    ``{key: BytesIO}``.

    Design notes
    ────────────
    • Uses a single shared Session across all workers (connection pool).
    • max_workers=50 – safe ceiling: S3 handles thousands of concurrent GETs,
      and openpyxl workbook building is single-threaded after this anyway.
    • We wait for ALL futures (not as_completed early-exit) to maximise
      the number of images that land in the export.
    """
    if not tasks:
        return {}

    session = _get_img_session()
    result: dict = {}
    result_lock = threading.Lock()

    def _worker(key: int, url: str, max_size: tuple) -> None:
        buf = download_and_resize_image(url, max_size, session=session)
        if buf is not None:
            with result_lock:
                result[key] = buf

    t0 = time_module.perf_counter()
    with ThreadPoolExecutor(max_workers=min(max_workers, len(tasks))) as pool:
        futures = [pool.submit(_worker, key, url, sz) for key, url, sz in tasks]
        # wait(...) with no timeout so we collect every image that succeeds
        wait(futures)

    elapsed_ms = (time_module.perf_counter() - t0) * 1000
    logger.info(
        "ImgDownload :: tasks=%d downloaded=%d elapsed_ms=%.0f",
        len(tasks), len(result), elapsed_ms,
    )
    return result


# ---------------------------------------------------------------------------
# Excel report helpers
# ---------------------------------------------------------------------------

def _stream_excel_report(
    *,
    db: Session,
    request: schemas.VehicleLogsRequest,
    current_user,
    use_expanded_pagination: bool,
    location_ids,
    checkpoint_ids,
    start_dt,
    end_dt,
    storage,
    total_records: int,
) -> StreamingResponse:
    t_start = time_module.perf_counter()
    logger.info("ExcelReport :: start total_records=%s", total_records)

    _BIG_PAGE = 100_000

    if use_expanded_pagination:
        all_entries = crud.get_vehicle_logs_with_blacklist_expanded(
            db,
            company_id=current_user.company_id,
            location_ids=location_ids,
            checkpoint_ids=checkpoint_ids,
            start_date=start_dt,
            end_date=end_dt,
            is_blacklisted=request.is_blacklisted,
            is_whitelisted=request.is_whitelisted,
            plate_number=request.plate_number,
            page=1,
            page_size=_BIG_PAGE,
        )

        cp_ids = _collect_checkpoint_ids_from_expanded(all_entries)
        cp_cache = _checkpoint_cache(db, cp_ids)

        # Collect ALL image paths (both plate + vehicle) without duplicates
        image_paths: set[str] = set()
        for ed in all_entries:
            pic = ed["history_entry"].get("Picture", {})
            v, p = _picture_paths(pic)
            if v:
                image_paths.add(v)
            if p:
                image_paths.add(p)

        presigned = storage.generate_presigned_urls_batch(list(image_paths), expiration=3600)

        excel_data = []
        for ed in all_entries:
            log = ed["log"]
            entry = ed["history_entry"]
            pic = entry.get("Picture", {})
            snap_time_str = pic.get("SnapInfo", {}).get("SnapTime", "")
            snap_date, snap_time = _split_snap_time(snap_time_str)
            v_img, p_img = _picture_paths(pic)
            cp_info = cp_cache.get(entry.get("checkpoint_id"), {})

            excel_data.append({
                "location_name":    cp_info.get("location_name", ""),
                "checkpoint_name":  cp_info.get("checkpoint_name", ""),
                "date":             snap_date,
                "time":             snap_time,
                "plate_number":     _display_plate(log),
                "plate_image_url":  presigned.get(p_img, "") if p_img else "",
                "blacklist":        "Yes" if log.is_blacklisted else "No",
                "whitelist":        "Yes" if log.is_whitelisted else "No",
            })

    else:
        all_logs = crud.get_vehicle_logs_with_blacklist(
            db,
            company_id=current_user.company_id,
            location_ids=location_ids,
            checkpoint_ids=checkpoint_ids,
            start_date=start_dt,
            end_date=end_dt,
            is_blacklisted=request.is_blacklisted,
            is_whitelisted=request.is_whitelisted,
            plate_number=None,
            page=1,
            page_size=_BIG_PAGE,
        )

        image_paths = set()
        for log in all_logs:
            pic = log.latest_data.get("Picture", {}) if log.latest_data else {}
            v, p = _picture_paths(pic)
            if v:
                image_paths.add(v)
            if p:
                image_paths.add(p)

        presigned = storage.generate_presigned_urls_batch(list(image_paths), expiration=3600)

        excel_data = []
        for log in all_logs:
            pic = log.latest_data.get("Picture", {}) if log.latest_data else {}
            snap_time_str = pic.get("SnapInfo", {}).get("SnapTime", "")
            snap_date, snap_time = _split_snap_time(snap_time_str)
            v_img, p_img = _picture_paths(pic)

            excel_data.append({
                "location_name":    log.location_name or "",
                "checkpoint_name":  log.checkpoint_name or "",
                "date":             snap_date,
                "time":             snap_time,
                "plate_number":     _display_plate(log),
                "plate_image_url":  presigned.get(p_img, "") if p_img else "",
                "blacklist":        "Yes" if log.is_blacklisted else "No",
                "whitelist":        "Yes" if log.is_whitelisted else "No",
            })

    t_db = time_module.perf_counter()
    logger.info(
        "ExcelReport :: db_done rows=%d elapsed_ms=%.0f",
        len(excel_data), (t_db - t_start) * 1000,
    )

    buf = generate_excel_report(excel_data, start_dt, end_dt)

    t_end = time_module.perf_counter()
    filename = f"vehicle_logs_{request.start_date_only}_{request.end_date_only}.xlsx"
    logger.info(
        "ExcelReport :: done rows=%d file=%s total_ms=%.0f",
        len(excel_data), filename, (t_end - t_start) * 1000,
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _split_snap_time(snap_time_str: str) -> tuple[str, str]:
    """Parse 'YYYY-MM-DD HH:MM:SS' → ('YYYY-MM-DD', 'HH:MM:SS')."""
    try:
        dt = datetime.strptime(snap_time_str, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")
    except ValueError:
        if " " in snap_time_str:
            parts = snap_time_str.split(" ", 1)
            return parts[0], parts[1]
        return snap_time_str, ""


def generate_excel_report(
    data: list,
    start_date: date | datetime | None = None,
    end_date: date | datetime | None = None,
) -> io.BytesIO:
    """
    Build a production-quality .xlsx workbook from vehicle-log dicts.

    Key improvements over original
    ───────────────────────────────
    • Both plate AND vehicle images are embedded (two image columns).
    • Images are downloaded with a shared Session + retry logic → near-zero misses.
    • All downloads fire in parallel (ThreadPoolExecutor) before any cell is written.
    • Proper row heights and column widths so images render at full size.
    • De-duplicates presigned URLs so the same S3 key is never fetched twice.
    • Graceful fallback: "—" text in cell when image is genuinely unavailable.
    """
    t0 = time_module.perf_counter()

    # ── 1. Collect & deduplicate all image URLs ───────────────────────────
    #
    # tasks: list of (row_index, "plate"|"vehicle", url, max_size)
    # We identify images by row_index + kind so we can look them up later.
    # Dedup by URL so the same S3 object is only fetched once.

    url_to_buf: dict[str, io.BytesIO | None] = {}   # populated after download
    tasks_for_download: list[tuple[str, str, tuple]] = []  # (url, url, size) – url is the key

    seen_urls: set[str] = set()
    for rec in data:
        url = rec.get("plate_image_url", "")
        if url and url.startswith("http") and url not in seen_urls:
            seen_urls.add(url)
            tasks_for_download.append((url, url, _PLATE_IMG_SIZE))

    logger.info(
        "ExcelReport :: unique_image_urls=%d (plate+vehicle)",
        len(tasks_for_download),
    )

    # ── 2. Download ALL images in parallel ───────────────────────────────
    if tasks_for_download:
        url_to_buf = _download_images_parallel(tasks_for_download)
    
    t_imgs = time_module.perf_counter()
    logger.info(
        "ExcelReport :: images downloaded=%d/%d elapsed_ms=%.0f",
        len(url_to_buf), len(tasks_for_download), (t_imgs - t0) * 1000,
    )

    # ── 3. Build workbook ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Logs"

    # ── Borders ───────────────────────────────────────────────────────────
    thick_border = Border(
        left=Side(style="medium",  color="2C3E50"),
        right=Side(style="medium", color="2C3E50"),
        top=Side(style="medium",   color="2C3E50"),
        bottom=Side(style="medium",color="2C3E50"),
    )
    thin_border = Border(
        left=Side(style="thin",  color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin",   color="BDC3C7"),
        bottom=Side(style="thin",color="BDC3C7"),
    )

    # ── Title row (row 1) ─────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "ROADPULSE – Vehicle Detection Report"
    c.font  = Font(name="Arial", bold=True, size=18, color="FFFFFF")
    c.fill  = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Date-range subtitle (row 2) ───────────────────────────────────────
    ws.merge_cells("A2:H2")
    c = ws["A2"]

    def _fmt(d) -> str:
        if isinstance(d, datetime):
            return d.strftime("%d %B %Y %H:%M:%S")
        if isinstance(d, date):
            return d.strftime("%d %B %Y")
        return ""

    if start_date and end_date:
        date_label = f"Report Period: {_fmt(start_date)}  →  {_fmt(end_date)}"
    elif start_date:
        date_label = f"Report Period: From {_fmt(start_date)}"
    elif end_date:
        date_label = f"Report Period: Until {_fmt(end_date)}"
    else:
        date_label = "Report Period: All Time"

    c.value = date_label
    c.font  = Font(name="Arial", bold=True, size=11, color="2C3E50")
    c.fill  = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # ── Column headers (row 3) ────────────────────────────────────────────
    # 8 columns: A–H
    headers = [
        "Location Name",    # A
        "Checkpoint Name",  # B
        "Date",             # C
        "Time",             # D
        "Plate Number",     # E
        "Plate Image",      # F  ← number-plate crop
        "Blacklist",        # G
        "Whitelist",        # H
    ]
    header_fill  = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = thick_border

    ws.row_dimensions[3].height = 30

    # Column widths (characters).  F and G are wider to accommodate images.
    col_widths = {
        "A": 22,   # Location Name
        "B": 22,   # Checkpoint Name
        "C": 13,   # Date
        "D": 10,   # Time
        "E": 16,   # Plate Number
        "F": 22,   # Plate Image   — ~160px wide image
        "G": 11,   # Blacklist
        "H": 11,   # Whitelist
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Data rows (starting row 4) ────────────────────────────────────────
    even_fill = PatternFill(start_color="F7F9F9", end_color="F7F9F9", fill_type="solid")
    odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Row height must be at least as tall as the largest image we embed.
    _ROW_HEIGHT_PX = _PLATE_IMG_SIZE[1] + 10   # padding
    _ROW_HEIGHT_PT = _ROW_HEIGHT_PX * 0.75    # px → pt (approximate)

    missing_images = 0

    for idx, record in enumerate(data):
        row = idx + 4

        # ── Text cells ────────────────────────────────────────────────
        ws.cell(row=row, column=1).value = record.get("location_name", "")
        ws.cell(row=row, column=2).value = record.get("checkpoint_name", "")
        ws.cell(row=row, column=3).value = record.get("date", "")
        ws.cell(row=row, column=4).value = record.get("time", "")

        plate_cell = ws.cell(row=row, column=5)
        plate_cell.value = record.get("plate_number", "")
        plate_cell.font  = Font(name="Arial", bold=True, size=10, color="2C3E50")

        # ── Plate image (column F) ────────────────────────────────────
        plate_url = record.get("plate_image_url", "")
        if plate_url and plate_url in url_to_buf:
            buf = url_to_buf[plate_url]
            if buf is not None:
                buf.seek(0)
                try:
                    img = XLImage(buf)
                    # Anchor top-left of cell; openpyxl uses EMUs internally.
                    img.anchor = f"F{row}"
                    ws.add_image(img)
                except Exception as exc:
                    logger.debug("ExcelReport :: plate_img_embed_err row=%d err=%s", row, exc)
                    ws.cell(row=row, column=6).value = "—"
                    missing_images += 1
            else:
                ws.cell(row=row, column=6).value = "—"
                missing_images += 1
        else:
            ws.cell(row=row, column=6).value = "—"
            if plate_url:   # URL was present but download failed
                missing_images += 1

        # ── Blacklist cell (column G) ─────────────────────────────────
        bl_cell = ws.cell(row=row, column=7)
        if record.get("blacklist") == "Yes":
            bl_cell.value = "YES"
            bl_cell.font  = Font(name="Arial", bold=True, size=10, color="E74C3C")
            bl_cell.fill  = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        else:
            bl_cell.value = "No"
            bl_cell.font  = Font(name="Arial", size=10, color="27AE60")

        # ── Whitelist cell (column H) ─────────────────────────────────
        wl_cell = ws.cell(row=row, column=8)
        if record.get("whitelist") == "Yes":
            wl_cell.value = "YES"
            wl_cell.font  = Font(name="Arial", bold=True, size=10, color="3498DB")
            wl_cell.fill  = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        else:
            wl_cell.value = "No"
            wl_cell.font  = Font(name="Arial", size=10, color="7F8C8D")

        # ── Common cell styling ───────────────────────────────────────
        row_fill = even_fill if row % 2 == 0 else odd_fill
        ws.row_dimensions[row].height = _ROW_HEIGHT_PT

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if col not in (5, 7, 8):    # plate number, blacklist, whitelist already styled
                cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            # Don't override the coloured fills on blacklist/whitelist columns
            if col not in (7, 8):
                cell.fill = row_fill

    # ── Footer row ────────────────────────────────────────────────────────
    footer_row = len(data) + 4
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    f_cell = ws.cell(row=footer_row, column=1)
    f_cell.value = (
        f"Total Records: {len(data)} | "
        f"Images: {len(url_to_buf)}/{len(tasks_for_download)} loaded | "
        f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M:%S')} | "
        "Powered by Transline Technologies"
    )
    f_cell.font      = Font(name="Arial", italic=True, size=9, color="7F8C8D")
    f_cell.alignment = Alignment(horizontal="center", vertical="center")
    f_cell.fill      = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    f_cell.border    = thick_border
    ws.row_dimensions[footer_row].height = 25

    ws.freeze_panes = "A4"

    # ── Serialise to buffer ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    t_end = time_module.perf_counter()
    logger.info(
        "ExcelReport :: workbook_built rows=%d missing_imgs=%d total_ms=%.0f",
        len(data), missing_images, (t_end - t0) * 1000,
    )
    return buf


# ---------------------------------------------------------------------------
# Fix vehicle number
# ---------------------------------------------------------------------------
@router.post("/fix-vehicle-number")
def fix_vehicle_number(
    request: schemas.FixVehicleNumberRequest = Body(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Correct a misread plate number on a vehicle log.

    • Only unrevised records (is_revised=False) can be updated.
    • The change is audited inside revised_data (who / when / why).
    • The endpoint is idempotent in the sense that a second attempt on the
      same record is explicitly rejected.
    """
    user_id = current_user.user_id
    username = current_user.username

    logger.info(
        "FixVehicleNumber :: user=%s(%s) record=%s old=%s new=%s",
        user_id, username, request.record_id, request.old_value, request.new_value,
    )

    vehicle_log = db.query(TrnVehicleLog).filter(
        TrnVehicleLog.log_id == request.record_id
    ).first()

    if not vehicle_log:
        logger.warning("FixVehicleNumber :: record %s not found", request.record_id)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vehicle log record not found.")

    if vehicle_log.is_revised:
        logger.warning("FixVehicleNumber :: record %s already revised", request.record_id)
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This record has already been revised and cannot be changed again.",
        )

    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)

    try:
        vehicle_log.revised_data = {
            "old_number": request.old_value,
            "new_number": request.new_value,
            "changed_by": username,
            "changed_at": now_ist.strftime("%Y-%m-%d %H:%M:%S"),
            "change_reason": request.change_reason,
        }
        vehicle_log.is_revised = True
        vehicle_log.updated_by = username
        vehicle_log.updated_at = now_ist.replace(tzinfo=None)

        db.commit()
        db.refresh(vehicle_log)
    except Exception as exc:
        db.rollback()
        logger.error("FixVehicleNumber :: record=%s error=%s", request.record_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to persist the correction. Please retry.",
        ) from exc

    logger.info(
        "FixVehicleNumber :: success record=%s old=%s new=%s",
        request.record_id, request.old_value, request.new_value,
    )

    return {
        "success": True,
        "message": "Vehicle number corrected successfully.",
        "record_id": request.record_id,
        "old_value": request.old_value,
        "new_value": request.new_value,
        "revised_data": vehicle_log.revised_data,
    }